from django.shortcuts import render, redirect
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, logout
from django.http import HttpResponse
from django.core.serializers.json import DjangoJSONEncoder
from .models import MilkSale, MilkCollection
from .forms import MilkSaleForm, MilkCollectionForm, LoginForm
import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal

import json
from datetime import date

def home(request):
    return render(request, 'index.html')

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)  # Use the LoginForm instead of AuthenticationForm
        if form.is_valid():
            # Log the user in
            login(request, form.get_user())
            return redirect('home')  # Redirect to the home page after successful login
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')  # Redirect to the home page after logout



class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super(DecimalEncoder, self).default(obj)

def milk_collection_graph(request):
    # Check if the user is authenticated
    if not request.user.is_authenticated:
        return redirect('login')  # Redirect to the login page if not authenticated

    data = MilkCollection.objects.values_list('date', 'total')

    # Convert date objects to string representations
    chart_data = [{'date': entry[0].strftime('%Y-%m-%d'), 'total': entry[1]} for entry in data]

    # Check if the request wants to download the Excel file
    if 'excel' in request.GET.get('format', ''):
        # Generate Excel file
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write header row
        worksheet.append(['Date', 'Total'])

        # Write data rows
        for entry in data:
            worksheet.append([entry[0], entry[1]])

        # Set column width for the 'Date' column
        worksheet.column_dimensions[get_column_letter(1)].width = 15  # Adjust the width as needed

        # Create a response with Excel file content type
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=milk_collection_data.xlsx'

        # Save workbook to response
        workbook.save(response)

        return response
    else:
        chart_data_json = json.dumps(chart_data, cls=DecimalEncoder)
        return render(request, 'milk_collection_graph.html', {'chart_data': chart_data_json})

def add_milk_collection(request):
    # Check if the user is authenticated
    if not request.user.is_authenticated:
        return redirect('login')  # Redirect to the login page if not authenticated

    if request.method == 'POST':
        form = MilkCollectionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('milk_collection_graph')  # Redirect to the graph page after successful submission
    else:
        form = MilkCollectionForm()

    return render(request, 'add_milk_collection.html', {'form': form})

def delete_milk_collection(request, collection_id):
    # Check if the user is authenticated
    if not request.user.is_authenticated:
        return redirect('login')  # Redirect to the login page if not authenticated

    milk_collection = MilkCollection.objects.get(collection_id=collection_id)
    milk_collection.delete()
    return redirect('milk_collection_graph')  # Redirect to the graph page after successful deletion

def milk_sale_graph(request):
    # Check if the user is authenticated
    if not request.user.is_authenticated:
        return redirect('login')  # Redirect to the login page if not authenticated

    data = MilkSale.objects.values_list('date', 'total')

    # Convert date objects to string representations
    chart_data = [{'date': entry[0].strftime('%Y-%m-%d'), 'total': entry[1]} for entry in data]

    # Convert the data to JSON
    chart_data_json = json.dumps(chart_data, cls=DjangoJSONEncoder)

    return render(request, 'milk_sale_graph.html', {'chart_data': chart_data_json})

def create_milk_sale(request):
    # Check if the user is authenticated
    if not request.user.is_authenticated:
        return redirect('login')  # Redirect to the login page if not authenticated

    if request.method == 'POST':
        form = MilkSaleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('milk_sale_graph')  # Redirect to the graph page after successful submission
    else:
        form = MilkSaleForm()

    return render(request, 'create_milk_sale.html', {'form': form})
